import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter

df = pd.read_csv('/home/claude/project/data/transactions_raw.csv')
df['transaction_datetime'] = pd.to_datetime(df['transaction_datetime'])
df['txn_hour'] = df['transaction_datetime'].dt.hour
df['txn_month'] = df['transaction_datetime'].dt.strftime('%Y-%m')
df['month_name'] = df['transaction_datetime'].dt.strftime('%b')
df['day_of_week'] = df['transaction_datetime'].dt.day_name()
df['amount_bucket'] = pd.cut(df['amount'], bins=[0,500,2000,5000,10000,50000], labels=['Micro (<500)','Small (500-2K)','Medium (2K-5K)','Large (5K-10K)','High Value (10K+)'])
df['time_slot'] = pd.cut(df['txn_hour'], bins=[-1,5,11,16,21,24], labels=['Night','Morning','Afternoon','Evening','Night2'])
df['time_slot'] = df['time_slot'].replace('Night2','Night')

wb = Workbook()

# --- Styles ---
header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill('solid', fgColor='2F5496')
subheader_fill = PatternFill('solid', fgColor='D6E4F0')
subheader_font = Font(name='Arial', bold=True, size=10, color='1F3864')
data_font = Font(name='Arial', size=10)
number_fmt = '#,##0'
decimal_fmt = '#,##0.00'
pct_fmt = '0.0%'
currency_fmt = '₹#,##0.00'
thin_border = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7')
)
green_fill = PatternFill('solid', fgColor='E2EFDA')
red_fill = PatternFill('solid', fgColor='FCE4EC')
yellow_fill = PatternFill('solid', fgColor='FFF9C4')

def write_header(ws, row, headers, col_start=1):
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=col_start+i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def write_data_row(ws, row, values, col_start=1, formats=None):
    for i, v in enumerate(values):
        cell = ws.cell(row=row, column=col_start+i, value=v)
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        if formats and i < len(formats) and formats[i]:
            cell.number_format = formats[i]

def set_col_widths(ws, widths):
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i+1)].width = w

def add_section_title(ws, row, title, merge_end=7):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=merge_end)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(name='Arial', bold=True, size=13, color='1F3864')
    cell.alignment = Alignment(horizontal='left')

# ==============================================================================
# SHEET 1: EXECUTIVE DASHBOARD
# ==============================================================================
ws1 = wb.active
ws1.title = "Executive Dashboard"
ws1.sheet_properties.tabColor = '2F5496'
set_col_widths(ws1, [20,18,18,18,18,18,18])

add_section_title(ws1, 1, "DIGITAL PAYMENT TRANSACTION ANALYSIS — EXECUTIVE SUMMARY")

# KPI Cards
kpi_row = 3
kpi_headers = ['KPI', 'Value']
write_header(ws1, kpi_row, kpi_headers)

total_txn = len(df)
total_success = len(df[df['status']=='Success'])
total_failed = len(df[df['status']=='Failed'])
total_pending = len(df[df['status']=='Pending'])
total_volume = df[df['status']=='Success']['amount'].sum()
avg_txn_value = df['amount'].mean()
unique_users = df['user_id'].nunique()

kpis = [
    ('Total Transactions', total_txn, number_fmt),
    ('Successful Transactions', total_success, number_fmt),
    ('Failed Transactions', total_failed, number_fmt),
    ('Pending Transactions', total_pending, number_fmt),
    ('Success Rate (%)', total_success / total_txn, pct_fmt),
    ('Failure Rate (%)', total_failed / total_txn, pct_fmt),
    ('Total Successful Volume (₹)', total_volume, currency_fmt),
    ('Average Transaction Value (₹)', avg_txn_value, currency_fmt),
    ('Unique Users', unique_users, number_fmt),
    ('Avg Transactions per User', total_txn / unique_users, '0.0'),
    ('Lost Revenue from Failures (₹)', df[df['status']=='Failed']['amount'].sum(), currency_fmt),
]

for i, (label, val, fmt) in enumerate(kpis):
    r = kpi_row + 1 + i
    ws1.cell(row=r, column=1, value=label).font = Font(name='Arial', bold=True, size=10)
    ws1.cell(row=r, column=1).border = thin_border
    c = ws1.cell(row=r, column=2, value=val)
    c.number_format = fmt
    c.font = Font(name='Arial', size=11, bold=True, color='2F5496')
    c.border = thin_border
    c.alignment = Alignment(horizontal='center')
    if label == 'Success Rate (%)':
        c.fill = green_fill
    elif label == 'Failure Rate (%)':
        c.fill = red_fill

# Payment Method Breakdown (right side)
pm_row = 3
pm_col = 4
ws1.cell(row=pm_row, column=pm_col, value='Payment Method Summary').font = Font(name='Arial', bold=True, size=12, color='1F3864')
write_header(ws1, pm_row+1, ['Payment Method', 'Transactions', 'Success Rate', 'Avg Amount (₹)'], pm_col)

method_stats = df.groupby('payment_method').agg(
    txn_count=('transaction_id','count'),
    success=('status', lambda x: (x=='Success').sum()),
    avg_amt=('amount','mean')
).reset_index()
method_stats['success_rate'] = method_stats['success'] / method_stats['txn_count']
method_stats = method_stats.sort_values('txn_count', ascending=False)

for i, row_data in method_stats.iterrows():
    r = pm_row + 2 + list(method_stats.index).index(i)
    write_data_row(ws1, r, [row_data['payment_method'], row_data['txn_count'], row_data['success_rate'], row_data['avg_amt']],
                   col_start=pm_col, formats=[None, number_fmt, pct_fmt, currency_fmt])


# ==============================================================================
# SHEET 2: SUCCESS RATE ANALYSIS
# ==============================================================================
ws2 = wb.create_sheet("Success Rate Analysis")
ws2.sheet_properties.tabColor = '548235'
set_col_widths(ws2, [20,15,15,15,15,18,18])

add_section_title(ws2, 1, "TRANSACTION SUCCESS RATE ANALYSIS")

# Monthly success rate table
monthly = df.groupby('txn_month').agg(
    total=('transaction_id','count'),
    success=('status', lambda x: (x=='Success').sum()),
    failed=('status', lambda x: (x=='Failed').sum()),
    volume=('amount','sum')
).reset_index()
monthly['success_rate'] = monthly['success'] / monthly['total']

headers = ['Month', 'Total Txn', 'Successful', 'Failed', 'Success Rate', 'Total Volume (₹)']
write_header(ws2, 3, headers)
for i, row_data in monthly.iterrows():
    r = 4 + i
    write_data_row(ws2, r, [row_data['txn_month'], row_data['total'], row_data['success'], row_data['failed'], row_data['success_rate'], row_data['volume']],
                   formats=[None, number_fmt, number_fmt, number_fmt, pct_fmt, currency_fmt])

# Line Chart: Monthly Success Rate
chart1 = LineChart()
chart1.title = "Monthly Success Rate Trend"
chart1.y_axis.title = "Success Rate (%)"
chart1.x_axis.title = "Month"
chart1.style = 10
chart1.width = 22
chart1.height = 14
cats = Reference(ws2, min_col=1, min_row=4, max_row=3+len(monthly))
data_ref = Reference(ws2, min_col=5, min_row=3, max_row=3+len(monthly))
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats)
chart1.series[0].graphicalProperties.line.width = 25000
ws2.add_chart(chart1, "A" + str(5 + len(monthly)))

# Success Rate by City
city_row = 5 + len(monthly) + 18
add_section_title(ws2, city_row, "Success Rate by City")
city_stats = df.groupby('city').agg(
    total=('transaction_id','count'),
    success=('status', lambda x: (x=='Success').sum()),
    avg_amt=('amount','mean')
).reset_index()
city_stats['success_rate'] = city_stats['success'] / city_stats['total']
city_stats = city_stats.sort_values('total', ascending=False)

write_header(ws2, city_row+1, ['City', 'Total Txn', 'Successful', 'Success Rate', 'Avg Amount (₹)'])
for i, row_data in city_stats.iterrows():
    r = city_row + 2 + list(city_stats.index).index(i)
    write_data_row(ws2, r, [row_data['city'], row_data['total'], row_data['success'], row_data['success_rate'], row_data['avg_amt']],
                   formats=[None, number_fmt, number_fmt, pct_fmt, currency_fmt])


# ==============================================================================
# SHEET 3: PEAK TRANSACTION TIMES
# ==============================================================================
ws3 = wb.create_sheet("Peak Times Analysis")
ws3.sheet_properties.tabColor = 'BF8F00'
set_col_widths(ws3, [15,15,18,18,15])

add_section_title(ws3, 1, "PEAK TRANSACTION TIMES ANALYSIS")

# Hourly distribution
hourly = df.groupby('txn_hour').agg(
    total=('transaction_id','count'),
    volume=('amount','sum'),
    avg_amt=('amount','mean'),
    success_rate=('status', lambda x: (x=='Success').mean())
).reset_index()

write_header(ws3, 3, ['Hour', 'Total Txn', 'Volume (₹)', 'Avg Amount (₹)', 'Success Rate'])
for i, row_data in hourly.iterrows():
    r = 4 + i
    hr_label = f"{int(row_data['txn_hour']):02d}:00"
    write_data_row(ws3, r, [hr_label, row_data['total'], row_data['volume'], row_data['avg_amt'], row_data['success_rate']],
                   formats=[None, number_fmt, currency_fmt, currency_fmt, pct_fmt])

# Bar Chart: Hourly Distribution
chart2 = BarChart()
chart2.type = "col"
chart2.title = "Transaction Volume by Hour"
chart2.y_axis.title = "Number of Transactions"
chart2.x_axis.title = "Hour of Day"
chart2.style = 10
chart2.width = 24
chart2.height = 14
cats2 = Reference(ws3, min_col=1, min_row=4, max_row=27)
data2 = Reference(ws3, min_col=2, min_row=3, max_row=27)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
chart2.series[0].graphicalProperties.solidFill = "2F5496"
ws3.add_chart(chart2, "A29")

# Day of Week Analysis
dow_row = 46
add_section_title(ws3, dow_row, "Day of Week Analysis")
day_order = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
daily = df.groupby('day_of_week').agg(
    total=('transaction_id','count'),
    volume=('amount','sum'),
    success_rate=('status', lambda x: (x=='Success').mean())
).reindex(day_order).reset_index()

write_header(ws3, dow_row+1, ['Day', 'Total Txn', 'Volume (₹)', 'Success Rate'])
for i, row_data in daily.iterrows():
    r = dow_row + 2 + i
    write_data_row(ws3, r, [row_data['day_of_week'], row_data['total'], row_data['volume'], row_data['success_rate']],
                   formats=[None, number_fmt, currency_fmt, pct_fmt])

# Time Slot Summary
ts_row = dow_row + 11
add_section_title(ws3, ts_row, "Time Slot Summary")
time_slots = df.groupby('time_slot').agg(
    total=('transaction_id','count'),
    volume=('amount','sum'),
    success_rate=('status', lambda x: (x=='Success').mean())
).reset_index()

write_header(ws3, ts_row+1, ['Time Slot', 'Total Txn', 'Volume (₹)', 'Success Rate'])
for i, row_data in time_slots.iterrows():
    r = ts_row + 2 + i
    write_data_row(ws3, r, [row_data['time_slot'], row_data['total'], row_data['volume'], row_data['success_rate']],
                   formats=[None, number_fmt, currency_fmt, pct_fmt])


# ==============================================================================
# SHEET 4: FAILED TRANSACTION ANALYSIS
# ==============================================================================
ws4 = wb.create_sheet("Failed Txn Analysis")
ws4.sheet_properties.tabColor = 'C00000'
set_col_widths(ws4, [25,15,15,18,18])

add_section_title(ws4, 1, "FAILED TRANSACTION ANALYSIS")

# Failure reasons breakdown
failed_df = df[df['status']=='Failed']
reasons = failed_df.groupby('failure_reason').agg(
    count=('transaction_id','count'),
    avg_amt=('amount','mean'),
    avg_proc=('processing_time_sec','mean')
).reset_index()
reasons['pct'] = reasons['count'] / reasons['count'].sum()
reasons = reasons.sort_values('count', ascending=False)

write_header(ws4, 3, ['Failure Reason', 'Count', '% of Failures', 'Avg Amount (₹)', 'Avg Processing Time (s)'])
for i, row_data in reasons.iterrows():
    r = 4 + list(reasons.index).index(i)
    write_data_row(ws4, r, [row_data['failure_reason'], row_data['count'], row_data['pct'], row_data['avg_amt'], row_data['avg_proc']],
                   formats=[None, number_fmt, pct_fmt, currency_fmt, decimal_fmt])

# Pie Chart: Failure Reasons
chart3 = PieChart()
chart3.title = "Failure Reason Distribution"
chart3.style = 10
chart3.width = 18
chart3.height = 14
pie_cats = Reference(ws4, min_col=1, min_row=4, max_row=3+len(reasons))
pie_data = Reference(ws4, min_col=2, min_row=3, max_row=3+len(reasons))
chart3.add_data(pie_data, titles_from_data=True)
chart3.set_categories(pie_cats)
chart3.dataLabels = DataLabelList()
chart3.dataLabels.showPercent = True
ws4.add_chart(chart3, "A" + str(5 + len(reasons)))

# Failure by Payment Method
fm_row = 5 + len(reasons) + 18
add_section_title(ws4, fm_row, "Failure Analysis by Payment Method")
method_fail = failed_df.groupby('payment_method').agg(
    failed_count=('transaction_id','count'),
    lost_revenue=('amount','sum'),
    avg_failed_amt=('amount','mean')
).reset_index()
method_fail = method_fail.sort_values('lost_revenue', ascending=False)

write_header(ws4, fm_row+1, ['Payment Method', 'Failed Count', 'Lost Revenue (₹)', 'Avg Failed Amount (₹)'])
for i, row_data in method_fail.iterrows():
    r = fm_row + 2 + list(method_fail.index).index(i)
    write_data_row(ws4, r, [row_data['payment_method'], row_data['failed_count'], row_data['lost_revenue'], row_data['avg_failed_amt']],
                   formats=[None, number_fmt, currency_fmt, currency_fmt])
    ws4.cell(row=r, column=1).fill = red_fill

# Revenue Impact
ri_row = fm_row + 9
add_section_title(ws4, ri_row, "Revenue Impact Summary")
total_lost = failed_df['amount'].sum()
ws4.cell(row=ri_row+1, column=1, value="Total Lost Revenue (₹)").font = Font(name='Arial', bold=True, size=11)
c = ws4.cell(row=ri_row+1, column=2, value=total_lost)
c.number_format = currency_fmt
c.font = Font(name='Arial', bold=True, size=12, color='C00000')
c.fill = red_fill

avg_lost = failed_df['amount'].mean()
ws4.cell(row=ri_row+2, column=1, value="Avg Failed Transaction (₹)").font = Font(name='Arial', bold=True, size=11)
c2 = ws4.cell(row=ri_row+2, column=2, value=avg_lost)
c2.number_format = currency_fmt
c2.font = Font(name='Arial', bold=True, size=11, color='C00000')


# ==============================================================================
# SHEET 5: USER ACTIVITY TRENDS
# ==============================================================================
ws5 = wb.create_sheet("User Activity Trends")
ws5.sheet_properties.tabColor = '7030A0'
set_col_widths(ws5, [22,15,18,18,15])

add_section_title(ws5, 1, "USER ACTIVITY TRENDS")

# User Segmentation
user_stats = df.groupby('user_id').agg(
    txn_count=('transaction_id','count'),
    total_spent=('amount', lambda x: x[df.loc[x.index, 'status']=='Success'].sum()),
    avg_amount=('amount','mean'),
    methods_used=('payment_method','nunique')
).reset_index()

def segment(cnt):
    if cnt >= 20: return 'Power User (20+)'
    elif cnt >= 10: return 'Regular User (10-19)'
    elif cnt >= 5: return 'Occasional User (5-9)'
    else: return 'Rare User (1-4)'

user_stats['segment'] = user_stats['txn_count'].apply(segment)
seg_summary = user_stats.groupby('segment').agg(
    user_count=('user_id','count'),
    avg_txn=('txn_count','mean'),
    avg_spend=('total_spent','mean')
).reset_index()

seg_order = ['Power User (20+)', 'Regular User (10-19)', 'Occasional User (5-9)', 'Rare User (1-4)']
seg_summary['segment'] = pd.Categorical(seg_summary['segment'], categories=seg_order, ordered=True)
seg_summary = seg_summary.sort_values('segment')

write_header(ws5, 3, ['User Segment', 'User Count', 'Avg Txn/User', 'Avg Spend (₹)'])
for i, row_data in seg_summary.iterrows():
    r = 4 + list(seg_summary.index).index(i)
    write_data_row(ws5, r, [row_data['segment'], row_data['user_count'], row_data['avg_txn'], row_data['avg_spend']],
                   formats=[None, number_fmt, '0.0', currency_fmt])

# Bar Chart: User Segments
chart4 = BarChart()
chart4.type = "col"
chart4.title = "User Segmentation"
chart4.y_axis.title = "Number of Users"
chart4.style = 10
chart4.width = 20
chart4.height = 12
cats4 = Reference(ws5, min_col=1, min_row=4, max_row=3+len(seg_summary))
data4 = Reference(ws5, min_col=2, min_row=3, max_row=3+len(seg_summary))
chart4.add_data(data4, titles_from_data=True)
chart4.set_categories(cats4)
chart4.series[0].graphicalProperties.solidFill = "7030A0"
ws5.add_chart(chart4, "A" + str(5 + len(seg_summary)))

# Monthly Active Users
mau_row = 5 + len(seg_summary) + 17
add_section_title(ws5, mau_row, "Monthly Active Users (MAU)")
mau = df.groupby('txn_month').agg(
    active_users=('user_id','nunique'),
    total_txn=('transaction_id','count'),
).reset_index()
mau['txn_per_user'] = mau['total_txn'] / mau['active_users']

write_header(ws5, mau_row+1, ['Month', 'Active Users', 'Total Txn', 'Txn/User'])
for i, row_data in mau.iterrows():
    r = mau_row + 2 + i
    write_data_row(ws5, r, [row_data['txn_month'], row_data['active_users'], row_data['total_txn'], row_data['txn_per_user']],
                   formats=[None, number_fmt, number_fmt, '0.0'])

# Top 15 Users
top_row = mau_row + 2 + len(mau) + 2
add_section_title(ws5, top_row, "Top 15 Users by Spend", merge_end=5)
top_users = user_stats.nlargest(15, 'total_spent')
write_header(ws5, top_row+1, ['User ID', 'Total Txn', 'Total Spent (₹)', 'Avg Amount (₹)', 'Methods Used'])
for i, (_, row_data) in enumerate(top_users.iterrows()):
    r = top_row + 2 + i
    write_data_row(ws5, r, [row_data['user_id'], row_data['txn_count'], row_data['total_spent'], row_data['avg_amount'], row_data['methods_used']],
                   formats=[None, number_fmt, currency_fmt, currency_fmt, number_fmt])


# ==============================================================================
# SHEET 6: RAW DATA
# ==============================================================================
ws6 = wb.create_sheet("Raw Data")
ws6.sheet_properties.tabColor = '808080'

raw_headers = ['Transaction ID', 'User ID', 'DateTime', 'Payment Method', 'Category', 'Merchant',
               'Amount (₹)', 'Status', 'Failure Reason', 'Platform', 'City', 'Processing Time (s)']
write_header(ws6, 1, raw_headers)

for col_idx in range(1, 13):
    ws6.column_dimensions[get_column_letter(col_idx)].width = 18

for i, (_, row_data) in enumerate(df.iterrows()):
    r = 2 + i
    ws6.cell(row=r, column=1, value=row_data['transaction_id']).font = data_font
    ws6.cell(row=r, column=2, value=row_data['user_id']).font = data_font
    ws6.cell(row=r, column=3, value=str(row_data['transaction_datetime'])).font = data_font
    ws6.cell(row=r, column=4, value=row_data['payment_method']).font = data_font
    ws6.cell(row=r, column=5, value=row_data['category']).font = data_font
    ws6.cell(row=r, column=6, value=row_data['merchant']).font = data_font
    c = ws6.cell(row=r, column=7, value=row_data['amount'])
    c.font = data_font
    c.number_format = currency_fmt
    ws6.cell(row=r, column=8, value=row_data['status']).font = data_font
    if row_data['status'] == 'Success':
        ws6.cell(row=r, column=8).fill = green_fill
    elif row_data['status'] == 'Failed':
        ws6.cell(row=r, column=8).fill = red_fill
    else:
        ws6.cell(row=r, column=8).fill = yellow_fill
    ws6.cell(row=r, column=9, value=row_data['failure_reason'] if pd.notna(row_data['failure_reason']) else '').font = data_font
    ws6.cell(row=r, column=10, value=row_data['platform']).font = data_font
    ws6.cell(row=r, column=11, value=row_data['city']).font = data_font
    c2 = ws6.cell(row=r, column=12, value=row_data['processing_time_sec'])
    c2.font = data_font
    c2.number_format = decimal_fmt

# Add autofilter
ws6.auto_filter.ref = f"A1:L{len(df)+1}"

# Freeze panes on raw data
ws6.freeze_panes = 'A2'

outpath = '/home/claude/project/excel/Payment_Transaction_Analysis.xlsx'
wb.save(outpath)
print(f"Excel workbook saved to {outpath}")
